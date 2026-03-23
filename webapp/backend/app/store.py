from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.schemas import OrderResponse, ProductResponse, UserResponse


USERS: dict[str, dict[str, str]] = {
    "admin@shop.local": {
        "password": "admin123",
        "full_name": "Никифорова Весения Николаевна",
        "role": "Администратор",
    },
    "manager@shop.local": {
        "password": "manager123",
        "full_name": "Степанов Михаил Артёмович",
        "role": "Менеджер",
    },
    "client@shop.local": {
        "password": "client123",
        "full_name": "Петрова Ольга Ивановна",
        "role": "Авторизованный клиент",
    },
}

PROJECT_ROOT = Path(__file__).resolve().parents[3]
IMPORT_DIR = PROJECT_ROOT / "import"


def _demo_products() -> dict[str, ProductResponse]:
    return {
        "A112T4": ProductResponse(
            article="A112T4",
            name="Ботинки",
            category="Женская обувь",
            description="Женские ботинки демисезонные kari",
            manufacturer="Kari",
            supplier="Kari",
            price=4990,
            unit="шт.",
            stock_qty=6,
            discount_percent=3,
            image_url="/images/1.jpg",
            final_price=4840.3,
        ),
        "F635R4": ProductResponse(
            article="F635R4",
            name="Туфли",
            category="Женская обувь",
            description="Туфли Marco Tozzi, размер 39",
            manufacturer="Marco Tozzi",
            supplier="ООО Марко",
            price=3244,
            unit="шт.",
            stock_qty=13,
            discount_percent=2,
            image_url="/images/2.jpg",
            final_price=3179.12,
        ),
    }


def _load_products_from_xlsx() -> dict[str, ProductResponse]:
    workbook_path = IMPORT_DIR / "Tovar.xlsx"
    if not workbook_path.exists():
        return _demo_products()

    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.worksheets[0]
    loaded: dict[str, ProductResponse] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        article = str(row[0]).strip()
        name = str(row[1] or "").strip()
        unit = str(row[2] or "шт.").strip()
        price = float(row[3] or 0)
        supplier = str(row[4] or "Не указан").strip()
        manufacturer = str(row[5] or "Не указан").strip()
        category = str(row[6] or "Без категории").strip()
        discount_percent = int(row[7] or 0)
        stock_qty = int(row[8] or 0)
        description = str(row[9] or "Без описания").strip()
        image_name = str(row[10] or "picture.png").strip()

        final_price = round(price * (1 - discount_percent / 100), 2)
        loaded[article] = ProductResponse(
            article=article,
            name=name,
            category=category,
            description=description,
            manufacturer=manufacturer,
            supplier=supplier,
            price=price,
            unit=unit,
            stock_qty=stock_qty,
            discount_percent=discount_percent,
            image_url=f"/images/{image_name}",
            final_price=final_price,
        )

    return loaded if loaded else _demo_products()


PRODUCTS: dict[str, ProductResponse] = _load_products_from_xlsx()

ORDERS: dict[int, OrderResponse] = {
    1: OrderResponse(
        id=1,
        status="Завершен",
        pickup_address="420151, г. Лесной, ул. Вишневая, 32",
        order_date=date(2025, 2, 27),
        delivery_date=date(2025, 4, 20),
        client_full_name="Степанов Михаил Артёмович",
        receive_code="901",
        items=[
            {"article": "A112T4", "quantity": 2},
            {"article": "F635R4", "quantity": 2},
        ],
    )
}


def authenticate(login: str, password: str) -> UserResponse | None:
    user = USERS.get(login)
    if user is None or user["password"] != password:
        return None
    return UserResponse(full_name=user["full_name"], role=user["role"])


def list_products() -> list[ProductResponse]:
    return list(PRODUCTS.values())


def put_product(article: str, payload: ProductResponse) -> ProductResponse:
    PRODUCTS[article] = payload
    return payload


def remove_product(article: str) -> bool:
    if any(article == item.article for order in ORDERS.values() for item in order.items):
        return False
    PRODUCTS.pop(article, None)
    return True


def list_orders() -> list[OrderResponse]:
    return list(ORDERS.values())


def put_order(order_id: int, payload: OrderResponse) -> OrderResponse:
    ORDERS[order_id] = payload
    return payload


def remove_order(order_id: int) -> None:
    ORDERS.pop(order_id, None)
